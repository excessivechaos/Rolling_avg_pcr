import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
import PySimpleGUI as sg
import os
import subprocess
import math
import platform


def calculate_pcr(df: pd.DataFrame) -> float:
    if df.columns[0] == "Date Opened":  # OO BT data
        return df["P/L"].sum() / (df["Premium"] * df["No. of Contracts"]).sum()
    elif df.columns[0] == "TradeID":  # BYOB BT data
        df["True P/L"] = df["ProfitLossAfterSlippage"] - df["CommissionFees"] / 100
        return df["True P/L"].sum() / df["Premium"].sum()
    else:
        raise ValueError("Unknown dataset type")


def analyze(file, short_avg_period, long_avg_period, short_weight, long_weight, top_x):
    # Load the CSV file
    df = pd.read_csv(file)

    # Determine which type of data, OptionOmega or BYOB
    if df.columns[0] == "Date Opened":  # OO BT data
        # Convert 'Date Opened' to datetime format
        df["Date Opened"] = pd.to_datetime(df["Date Opened"])

        # Sort by 'Date Opened' and 'Time Opened'
        df.sort_values(["Date Opened", "Time Opened"], inplace=True)

        # Group by month and 'Time Opened'
        df_grouped = df.groupby([df["Date Opened"].dt.to_period("M"), "Time Opened"])

        # Determine start and end dates
        start_date = df["Date Opened"].min().date()
        end_date = df["Date Opened"].max().date()

    elif df.columns[0] == "TradeID":  # BYOB BT data
        # Convert 'EntryTime' to datetime format
        df["EntryTime"] = pd.to_datetime(df["EntryTime"])

        # Create a 'Time' column
        df["Time"] = df["EntryTime"].dt.strftime("%H:%M")

        # Sort by 'EntryTime'
        df.sort_values(["EntryTime"], inplace=True)

        # Group by month and time
        df_grouped = df.groupby([df["EntryTime"].dt.to_period("M"), df["Time"]])

        # Determine start and end dates
        start_date = df["EntryTime"].min().date()
        end_date = df["EntryTime"].max().date()

    else:
        raise ValueError("Unknown dataset type")

    # Calculate PCR for each group
    df_pcr = df_grouped.apply(calculate_pcr)

    # Unstack if it's a MultiIndex
    if isinstance(df_pcr.index, pd.MultiIndex):
        df_pcr = df_pcr.unstack()

    # Calculate rolling averages
    df_pcr_short_avg = df_pcr.rolling(short_avg_period, min_periods=1).mean()
    df_pcr_long_avg = df_pcr.rolling(long_avg_period, min_periods=1).mean()

    # Calculate weighted average
    df_pcr_weighted_avg = short_weight * df_pcr_short_avg + long_weight * df_pcr_long_avg

    # Calculate 1-month average PCR
    df_pcr_1mo_avg = df_pcr.rolling(1, min_periods=1).mean()

    # Sort the data in descending order by date
    df_pcr_weighted_avg.sort_index(ascending=False, inplace=True)
    df_pcr_1mo_avg.sort_index(ascending=False, inplace=True)

    # Convert df_pcr_weighted_avg to a DataFrame if it's a Series
    if isinstance(df_pcr_weighted_avg, pd.Series):
        df_pcr_weighted_avg = df_pcr_weighted_avg.to_frame()

    # Convert PCR to percentage and round to 1 decimal place
    df_pcr_weighted_avg = df_pcr_weighted_avg.applymap(lambda x: round(x * 100, 1))
    df_pcr_1mo_avg = df_pcr_1mo_avg.applymap(lambda x: round(x * 100, 1))

    # path
    path = os.path.dirname(file)
    # Create filename
    filename = os.path.join(
        path,
        f"Weighted_trail_avg_pcr_{short_avg_period}mo-{long_avg_period}mo_{start_date}-{end_date}.xlsx",
    )

    # Create a new DataFrame for output, adding date range as index
    df_output = pd.DataFrame(index=df_pcr_weighted_avg.index)
    for i, (date, row) in enumerate(df_pcr_weighted_avg.iterrows()):
        current_month_end = date.to_timestamp() + pd.offsets.MonthEnd(1)
        previous_month_start = (
            current_month_end - pd.DateOffset(months=long_avg_period - 1)
        ).replace(day=1)
        if i == 0:
            date_range_label = f"{end_date} - {previous_month_start.date()}"
        elif i == len(df_pcr_weighted_avg) - 1:
            date_range_label = f"{current_month_end.date()} - {start_date}"
        else:
            date_range_label = (
                f"{current_month_end.date()} - {previous_month_start.date()}"
            )
        df_output.loc[date, "Date Range"] = date_range_label

    df_output_1mo_avg = pd.DataFrame(index=df_pcr_1mo_avg.index)
    for i, (date, row) in enumerate(df_pcr_1mo_avg.iterrows()):
        current_month_end = date.to_timestamp() + pd.offsets.MonthEnd(1)
        previous_month_start = (current_month_end - pd.DateOffset(months=1 - 1)).replace(
            day=1
        )
        if i == 0:
            date_range_label = f"{end_date} - {previous_month_start.date()}"
        elif i == len(df_pcr_1mo_avg) - 1:
            date_range_label = f"{current_month_end.date()} - {start_date}"
        else:
            date_range_label = (
                f"{current_month_end.date()} - {previous_month_start.date()}"
            )
        df_output_1mo_avg.loc[date, "Date Range"] = date_range_label

    # Concatenate the output DataFrame and the weighted PCR DataFrame
    df_output = pd.concat([df_output, df_pcr_weighted_avg], axis=1)

    # Concatenate the output DataFrame and the 1mo PCR DataFrame
    df_output_1mo_avg = pd.concat([df_output_1mo_avg, df_pcr_1mo_avg], axis=1)

    # Create a Pandas Excel writer using XlsxWriter as the engine
    with pd.ExcelWriter(filename, engine="xlsxwriter") as writer:
        # Write the DataFrame to an Excel file
        df_output.to_excel(writer, sheet_name="TW_PCR", index=False)
        df_output_1mo_avg.to_excel(writer, sheet_name="1mo Avg PCR", index=False)

        # Get the xlsxwriter workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets["TW_PCR"]
        worksheet_1mo = writer.sheets["1mo Avg PCR"]

        # Set the PCR columns to percentage format
        percent_format = workbook.add_format({"num_format": "0.0%"})
        top_x_format = workbook.add_format({"bold": 1, "font_color": "#FFFFFF"})  # white
        for row in range(
            2, len(df_output) + 2
        ):  # +2 because Excel's index starts from 1 and there is a header row
            # Apply a conditional format to the PCR cells in the current row
            worksheet.conditional_format(
                f"B{row}:{get_column_letter(len(df_output.columns))}{row}",
                {
                    "type": "3_color_scale",
                    "min_color": "red",
                    "mid_color": "yellow",
                    "max_color": "green",
                },
            )
            # Format top x values in bold white text
            if top_x > 0:
                worksheet.conditional_format(
                    f"B{row}:{get_column_letter(len(df_output.columns))}{row}",
                    {
                        "type": "top",
                        "value": top_x,
                        "format": top_x_format,
                    },
                )

            # Set the number format of the PCR cells in the current row to percentage
            for col in range(
                2, len(df_output.columns) + 2
            ):  # +2 because Excel's index starts from 1
                cell_value = df_output.iloc[row - 2, col - 2]

                if (
                    isinstance(cell_value, (int, float))
                    and not math.isnan(cell_value)
                    and not math.isinf(cell_value)
                ):
                    worksheet.write(row - 1, col - 2, cell_value / 100, percent_format)

        for row in range(
            2, len(df_pcr_1mo_avg) + 2
        ):  # +2 because Excel's index starts from 1 and there is a header row
            # Apply a conditional format to the PCR cells in the current row
            worksheet_1mo.conditional_format(
                f"B{row}:{get_column_letter(len(df_output_1mo_avg.columns))}{row}",
                {
                    "type": "3_color_scale",
                    "min_color": "red",
                    "mid_color": "yellow",
                    "max_color": "green",
                },
            )
            # Format top x values in bold white text
            if top_x > 0:
                worksheet_1mo.conditional_format(
                    f"B{row}:{get_column_letter(len(df_output_1mo_avg.columns))}{row}",
                    {
                        "type": "top",
                        "value": top_x,
                        "format": top_x_format,
                    },
                )

            # Set the number format of the PCR cells in the current row to percentage
            for col in range(
                2, len(df_output_1mo_avg.columns) + 2
            ):  # +2 because Excel's index starts from 1
                cell_value = df_output_1mo_avg.iloc[row - 2, col - 2]

                if (
                    isinstance(cell_value, (int, float))
                    and not math.isnan(cell_value)
                    and not math.isinf(cell_value)
                ):
                    worksheet_1mo.write(
                        row - 1, col - 2, cell_value / 100, percent_format
                    )

        # Adjust the column widths
        for column in df_output:
            column_length = max(
                df_output[column].astype(str).map(len).max() + 1, len(column) + 1
            )
            col_idx = df_output.columns.get_loc(column)
            worksheet.set_column(col_idx, col_idx, column_length)

        # Adjust the column widths
        for column in df_output_1mo_avg:
            column_length = max(
                df_output_1mo_avg[column].astype(str).map(len).max() + 1, len(column) + 1
            )
            col_idx = df_output_1mo_avg.columns.get_loc(column)
            worksheet_1mo.set_column(col_idx, col_idx, column_length)

    # open file in excel
    # subprocess.Popen(["open", filename], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if platform.system() == "Windows":
        subprocess.Popen(["cmd", "/c", "start", filename], shell=True)
    elif platform.system() == "Darwin":  # This is the value returned for macOS
        subprocess.Popen(["open", filename])
    else:
        print("Unsupported platform: ", platform.system())


def main():
    sg.theme("DarkGrey14")
    sg.SetOptions(font=("Arial", 12))

    layout = [
        [sg.Text("Select trade log CSV file:")],
        [sg.Input(key="-FILE-"), sg.FileBrowse()],
        [
            sg.Frame(
                "",
                [
                    [
                        sg.Text("Trailing Avg 1:"),
                        sg.Input(
                            "3", key="-AVG_PERIOD_1-", size=(3, 1), justification="c"
                        ),
                        sg.Text("Weight:"),
                        sg.Input(
                            "75", key="-PERIOD_1_WEIGHT-", size=(3, 1), justification="c"
                        ),
                    ],
                    [
                        sg.Text("Trailing Avg 2:"),
                        sg.Input(
                            "10", key="-AVG_PERIOD_2-", size=(3, 1), justification="c"
                        ),
                        sg.Text("Weight:"),
                        sg.Input(
                            "25", key="-PERIOD_2_WEIGHT-", size=(3, 1), justification="c"
                        ),
                    ],
                    [
                        sg.Text("Highlight Top", pad=(5, 5)),
                        sg.Input(
                            "5", key="-TOP_X-", size=(2, 1), pad=(0, 0), justification="c"
                        ),
                        sg.Text("Values with White Text", pad=(5, 0)),
                    ],
                    [
                        sg.Button("Analyze", pad=(5, 10)),
                        sg.Button("Cancel"),
                        sg.Text("", size=(30, 1)),
                        sg.Text("v1.0.1"),
                    ],
                ],
            )
        ],
    ]

    window = sg.Window("Rolling Avg PCR Analyzer", layout, resizable=True)

    while True:
        event, values = window.read()

        if event == "Cancel" or event == sg.WIN_CLOSED:
            break

        if event == "Analyze":
            if values["-FILE-"][-3:].lower() != "csv":
                sg.popup_no_border("Please select a csv file")
            else:
                analyze(
                    values["-FILE-"],
                    int(values["-AVG_PERIOD_1-"]),
                    int(values["-AVG_PERIOD_2-"]),
                    float(values["-PERIOD_1_WEIGHT-"]) / 100,
                    float(values["-PERIOD_2_WEIGHT-"]) / 100,
                    int(values["-TOP_X-"]),
                )

    window.close()


if __name__ == "__main__":
    main()